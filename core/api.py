from collections import defaultdict
from decimal import Decimal

from asgiref.sync import sync_to_async
from io import BytesIO

from adrf.views import APIView
from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.db.models import Count, Sum
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import exceptions, status
from rest_framework.response import Response

from core.models import Car, CarPhoto, DefectPhoto, Expense, ExpenseCategory, ExpensePhoto, TelegramUser
from core.serializers import (
    CarInputSerializer,
    CarPatchSerializer,
    CarStageSerializer,
    CarStatusSerializer,
    DefectPhotoInputSerializer,
    ExpenseInputSerializer,
    ExpensePatchSerializer,
    TelegramUserInputSerializer,
    serialize_car,
    serialize_defect_photo,
    serialize_expense,
)


def _api_key_allowed(request) -> bool:
    if not settings.INTERNAL_API_KEY:
        return True
    return request.headers.get("X-API-Key") == settings.INTERNAL_API_KEY


def _photo_items(items) -> list[dict[str, str]]:
    photos = []
    for item in items or []:
        if not isinstance(item, dict):
            continue
        file_id = str(item.get("file_id") or item.get("photo_file_id") or "")
        image_path = str(item.get("image_path") or item.get("path") or "")
        if file_id or image_path:
            photos.append({"file_id": file_id, "image_path": image_path})
    return photos


async def _get_category(data: dict):
    category_id = data.get("category_id")
    category_name = data.get("category_name", "").strip()
    if category_id:
        return await ExpenseCategory.objects.aget(id=category_id, is_active=True)
    if category_name:
        category, _ = await ExpenseCategory.objects.aget_or_create(name=category_name)
        return category
    return None


async def _get_user_by_telegram_id(telegram_id):
    if not telegram_id:
        return None
    try:
        return await TelegramUser.objects.aget(telegram_id=telegram_id, is_active=True)
    except TelegramUser.DoesNotExist:
        return None


def _serialize_user(user: TelegramUser) -> dict:
    return {
        "id": user.id,
        "telegram_id": user.telegram_id,
        "username": user.username,
        "full_name": user.full_name,
        "role": user.role,
        "is_active": user.is_active,
    }


@sync_to_async
def _create_car_with_photos_sync(data: dict, created_by, photos: list[dict[str, str]]) -> int:
    main_photo = photos[0] if photos else {}
    with transaction.atomic():
        car = Car.objects.create(
            title=data["title"],
            brand=data.get("brand", ""),
            model=data.get("model", ""),
            vin_or_plate=data.get("vin_or_plate", ""),
            vin=data.get("vin", ""),
            status=data.get("status", Car.Status.IN_WORK),
            repair_stage=data.get("repair_stage", Car.RepairStage.ACCEPTED),
            description=data.get("description", ""),
            car_photo_file_id=main_photo.get("file_id", ""),
            car_photo=main_photo.get("image_path", ""),
            vin_photo_file_id=data.get("vin_photo_file_id", ""),
            vin_photo=data.get("vin_photo_path", ""),
            created_by=created_by,
        )
        CarPhoto.objects.bulk_create(
            [
                CarPhoto(
                    car=car,
                    photo_file_id=photo.get("file_id", ""),
                    image=photo.get("image_path", ""),
                    created_by=created_by,
                )
                for photo in photos
            ]
        )
        return car.id


@sync_to_async
def _create_defect_photos_sync(car: Car, created_by, photos: list[dict[str, str]], comment: str = "") -> list[int]:
    with transaction.atomic():
        created = DefectPhoto.objects.bulk_create(
            [
                DefectPhoto(
                    car=car,
                    photo_file_id=photo.get("file_id", ""),
                    image=photo.get("image_path", ""),
                    comment=comment,
                    created_by=created_by,
                )
                for photo in photos
            ]
        )
        return [photo.id for photo in created]


@sync_to_async
def _create_expense_with_photos_sync(data: dict, car: Car, category, employee, photos: list[dict[str, str]]) -> int:
    main_photo = photos[0] if photos else {}
    with transaction.atomic():
        expense = Expense.objects.create(
            car=car,
            amount=data["amount"],
            currency=data.get("currency") or "BYN",
            description=data["description"],
            category=category,
            employee=employee,
            comment=data.get("comment", ""),
            receipt_photo_file_id=main_photo.get("file_id", ""),
            receipt_photo=main_photo.get("image_path", ""),
        )
        ExpensePhoto.objects.bulk_create(
            [
                ExpensePhoto(
                    expense=expense,
                    photo_file_id=photo.get("file_id", ""),
                    image=photo.get("image_path", ""),
                )
                for photo in photos
            ]
        )
        return expense.id


@sync_to_async
def _create_expenses_bulk_sync(items: list[dict], car: Car, category, employee, currency: str) -> list[int]:
    with transaction.atomic():
        locked_car = Car.objects.select_for_update().get(pk=car.pk)
        if locked_car.status != Car.Status.IN_WORK:
            raise ValueError("inactive_car")
        created = Expense.objects.bulk_create(
            [
                Expense(
                    car=locked_car,
                    amount=item["amount"],
                    currency=currency or "BYN",
                    description=item["description"],
                    category=category,
                    employee=employee,
                )
                for item in items
            ]
        )
        return [expense.id for expense in created]


class InternalAPIView(APIView):
    def initial(self, request, *args, **kwargs):
        super().initial(request, *args, **kwargs)
        if not _api_key_allowed(request):
            raise exceptions.PermissionDenied("Invalid API key.")


class TelegramUserView(InternalAPIView):
    async def post(self, request):
        serializer = TelegramUserInputSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        data = serializer.validated_data
        user, _ = await TelegramUser.objects.aupdate_or_create(
            telegram_id=data["telegram_id"],
            defaults={"username": data.get("username", ""), "full_name": data.get("full_name", "")},
        )
        return Response(_serialize_user(user))


class TelegramUserAccessView(InternalAPIView):
    async def get(self, request):
        telegram_id = request.query_params.get("telegram_id")
        if not telegram_id:
            return Response({"allowed": False, "detail": "telegram_id is required"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            user = await TelegramUser.objects.aget(telegram_id=telegram_id)
        except TelegramUser.DoesNotExist:
            return Response({"allowed": False, "telegram_id": int(telegram_id)})
        data = _serialize_user(user)
        data["allowed"] = user.is_active
        return Response(data)


class CategoryListView(InternalAPIView):
    async def get(self, request):
        categories = []
        async for category in ExpenseCategory.objects.filter(is_active=True).order_by("name"):
            categories.append({"id": category.id, "name": category.name})
        return Response(categories)


class CarListCreateView(InternalAPIView):
    async def get(self, request):
        status_filter = request.query_params.get("status")
        cars_qs = Car.objects.select_related("created_by").all()
        if status_filter:
            allowed_statuses = {choice[0] for choice in Car.Status.choices}
            if status_filter not in allowed_statuses:
                return Response({"detail": "Invalid status."}, status=status.HTTP_400_BAD_REQUEST)
            cars_qs = cars_qs.filter(status=status_filter)
        cars = [car async for car in cars_qs.order_by("status", "-created_at")]
        totals = defaultdict(lambda: 0)
        totals_by_currency = defaultdict(dict)
        counts = defaultdict(lambda: 0)
        if cars:
            async for row in Expense.objects.filter(car_id__in=[car.id for car in cars]).values("car_id", "currency").annotate(total=Sum("amount"), count=Count("id")):
                totals[row["car_id"]] += row["total"] or 0
                totals_by_currency[row["car_id"]][row["currency"] or "BYN"] = row["total"] or 0
                counts[row["car_id"]] += row["count"] or 0
        return Response([serialize_car(car, total=totals[car.id], expenses_count=counts[car.id], total_by_currency=totals_by_currency[car.id]) for car in cars])

    async def post(self, request):
        serializer = CarInputSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        data = serializer.validated_data
        created_by = await _get_user_by_telegram_id(data.get("created_by_telegram_id"))
        photos = _photo_items(data.get("photos"))
        if not photos and (data.get("car_photo_file_id") or data.get("car_photo_path")):
            photos.append({"file_id": data.get("car_photo_file_id", ""), "image_path": data.get("car_photo_path", "")})
        car_id = await _create_car_with_photos_sync(data, created_by, photos)
        car = await Car.objects.select_related("created_by").aget(pk=car_id)
        return Response(serialize_car(car, total=0), status=status.HTTP_201_CREATED)


class CarDetailView(InternalAPIView):
    async def get(self, request, pk: int):
        try:
            car = await Car.objects.select_related("created_by").aget(pk=pk)
        except Car.DoesNotExist:
            return Response({"detail": "Car not found."}, status=status.HTTP_404_NOT_FOUND)
        aggregate = await Expense.objects.filter(car=car).aaggregate(total=Sum("amount"))
        totals_by_currency = {}
        async for row in Expense.objects.filter(car=car).values("currency").annotate(total=Sum("amount")):
            totals_by_currency[row["currency"] or "BYN"] = row["total"] or 0
        return Response(serialize_car(car, total=aggregate["total"] or 0, total_by_currency=totals_by_currency))

    async def patch(self, request, pk: int):
        serializer = CarPatchSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        try:
            car = await Car.objects.select_related("created_by").aget(pk=pk)
        except Car.DoesNotExist:
            return Response({"detail": "Car not found."}, status=status.HTTP_404_NOT_FOUND)
        data = serializer.validated_data
        for field in ("title", "brand", "model", "vin_or_plate", "vin", "description", "car_photo_file_id", "vin_photo_file_id", "status", "repair_stage"):
            if field in data:
                setattr(car, field, data[field])
        if "car_photo_path" in data:
            car.car_photo = data["car_photo_path"]
        if "vin_photo_path" in data:
            car.vin_photo = data["vin_photo_path"]
        car.apply_status_dates()
        await car.asave()
        aggregate = await Expense.objects.filter(car=car).aaggregate(total=Sum("amount"), count=Count("id"))
        totals_by_currency = {}
        async for row in Expense.objects.filter(car=car).values("currency").annotate(total=Sum("amount")):
            totals_by_currency[row["currency"] or "BYN"] = row["total"] or 0
        return Response(serialize_car(car, total=aggregate["total"] or 0, expenses_count=aggregate["count"] or 0, total_by_currency=totals_by_currency))


class DefectPhotoListCreateView(InternalAPIView):
    async def get(self, request):
        car_id = request.query_params.get("car_id")
        photos_qs = DefectPhoto.objects.select_related("car", "created_by").all()
        if car_id:
            photos_qs = photos_qs.filter(car_id=car_id)
        photos = [serialize_defect_photo(photo) async for photo in photos_qs.order_by("-created_at")]
        return Response(photos)

    async def post(self, request):
        serializer = DefectPhotoInputSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        data = serializer.validated_data
        try:
            car = await Car.objects.aget(pk=data["car_id"])
        except Car.DoesNotExist:
            return Response({"detail": "Car not found."}, status=status.HTTP_404_NOT_FOUND)
        created_by = await _get_user_by_telegram_id(data.get("created_by_telegram_id"))
        photos = _photo_items(data.get("photos"))
        if not photos and (data.get("photo_file_id") or data.get("image_path")):
            photos.append({"file_id": data.get("photo_file_id", ""), "image_path": data.get("image_path", "")})
        if not photos:
            return Response({"detail": "At least one photo is required."}, status=status.HTTP_400_BAD_REQUEST)
        photo_ids = await _create_defect_photos_sync(car, created_by, photos, data.get("comment", ""))
        created = [photo async for photo in DefectPhoto.objects.select_related("car", "created_by").filter(id__in=photo_ids).order_by("id")]
        payload = [serialize_defect_photo(photo) for photo in created]
        if len(payload) == 1:
            return Response(payload[0], status=status.HTTP_201_CREATED)
        return Response(payload, status=status.HTTP_201_CREATED)


class ExpenseListCreateView(InternalAPIView):
    async def get(self, request):
        car_id = request.query_params.get("car_id")
        employee_telegram_id = request.query_params.get("employee_telegram_id")
        expenses_qs = Expense.objects.select_related("car", "category", "employee", "updated_by").all()
        if car_id:
            expenses_qs = expenses_qs.filter(car_id=car_id)
        if employee_telegram_id:
            expenses_qs = expenses_qs.filter(employee__telegram_id=employee_telegram_id)
        expenses = [serialize_expense(expense) async for expense in expenses_qs.order_by("-spent_at", "-created_at")]
        return Response(expenses)

    async def post(self, request):
        serializer = ExpenseInputSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        data = serializer.validated_data
        try:
            car = await Car.objects.aget(pk=data["car_id"])
            if car.status != Car.Status.IN_WORK:
                return Response({"detail": "Expenses can be added only to cars in work."}, status=status.HTTP_400_BAD_REQUEST)
            category = await _get_category(data)
        except ObjectDoesNotExist:
            return Response({"detail": "Related object not found."}, status=status.HTTP_404_NOT_FOUND)
        employee = await _get_user_by_telegram_id(data.get("employee_telegram_id"))
        photos = _photo_items(data.get("photos"))
        if not photos and (data.get("receipt_photo_file_id") or data.get("receipt_photo_path")):
            photos.append({"file_id": data.get("receipt_photo_file_id", ""), "image_path": data.get("receipt_photo_path", "")})
        expense_id = await _create_expense_with_photos_sync(data, car, category, employee, photos)
        expense = await Expense.objects.select_related("car", "category", "employee").aget(pk=expense_id)
        return Response(serialize_expense(expense), status=status.HTTP_201_CREATED)


class ExpenseBulkCreateView(InternalAPIView):
    async def post(self, request):
        data = dict(request.data)
        raw_items = data.get("items") or []
        if not isinstance(raw_items, list) or not raw_items:
            return Response({"items": "Добавьте хотя бы одну позицию расхода."}, status=status.HTTP_400_BAD_REQUEST)

        items = []
        for item in raw_items:
            payload = {
                "car_id": data.get("car_id"),
                "amount": item.get("amount") if isinstance(item, dict) else None,
                "currency": data.get("currency") or "BYN",
                "category_id": data.get("category_id"),
                "category_name": data.get("category_name", ""),
                "description": item.get("description") if isinstance(item, dict) else "",
                "employee_telegram_id": data.get("employee_telegram_id"),
            }
            serializer = ExpenseInputSerializer(data=payload)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            items.append(serializer.validated_data)

        first = items[0]
        try:
            car = await Car.objects.aget(pk=first["car_id"])
        except Car.DoesNotExist:
            return Response({"detail": "Car not found."}, status=status.HTTP_404_NOT_FOUND)
        try:
            category = await _get_category(first)
        except ExpenseCategory.DoesNotExist:
            return Response({"detail": "Category not found."}, status=status.HTTP_400_BAD_REQUEST)
        employee = await _get_user_by_telegram_id(first.get("employee_telegram_id"))

        try:
            ids = await _create_expenses_bulk_sync(items, car, category, employee, first.get("currency") or "BYN")
        except ValueError:
            return Response({"detail": "Cannot add expenses to completed or archived car."}, status=status.HTTP_400_BAD_REQUEST)

        expenses = [
            expense
            async for expense in Expense.objects.select_related("car", "category", "employee", "updated_by")
            .filter(id__in=ids)
            .order_by("id")
        ]
        return Response([serialize_expense(expense) for expense in expenses], status=status.HTTP_201_CREATED)


class ExpenseDetailView(InternalAPIView):
    async def patch(self, request, pk: int):
        serializer = ExpensePatchSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        try:
            expense = await Expense.objects.aget(pk=pk)
            category = await _get_category(serializer.validated_data)
        except ObjectDoesNotExist:
            return Response({"detail": "Related object not found."}, status=status.HTTP_404_NOT_FOUND)
        data = serializer.validated_data
        for field in ("amount", "currency", "description", "comment", "receipt_photo_file_id"):
            if field in data:
                setattr(expense, field, data[field])
        if "receipt_photo_path" in data:
            expense.receipt_photo = data["receipt_photo_path"]
        if data.get("updated_by_telegram_id"):
            expense.updated_by = await _get_user_by_telegram_id(data.get("updated_by_telegram_id"))
        if "category_id" in data or data.get("category_name"):
            expense.category = category
        await expense.asave()
        expense = await Expense.objects.select_related("car", "category", "employee", "updated_by").aget(pk=expense.pk)
        return Response(serialize_expense(expense))

    async def delete(self, request, pk: int):
        try:
            expense = await Expense.objects.aget(pk=pk)
        except Expense.DoesNotExist:
            return Response({"detail": "Expense not found."}, status=status.HTTP_404_NOT_FOUND)
        await expense.adelete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ReportSummaryView(InternalAPIView):
    async def get(self, request):
        rows = []
        cars = [car async for car in Car.objects.select_related("created_by").all().order_by("status", "-created_at")]
        totals = defaultdict(lambda: 0)
        totals_by_currency = defaultdict(dict)
        if cars:
            async for row in Expense.objects.filter(car_id__in=[car.id for car in cars]).values("car_id", "currency").annotate(total=Sum("amount")):
                totals[row["car_id"]] += row["total"] or 0
                totals_by_currency[row["car_id"]][row["currency"] or "BYN"] = row["total"] or 0
        for car in cars:
            rows.append(
                {
                    "car_id": car.id,
                    "title": car.title,
                    "make_model": car.vehicle_label,
                    "vin": car.vin_or_plate or car.vin,
                    "status": car.status,
                    "status_display": car.get_status_display(),
                    "repair_stage": car.repair_stage,
                    "repair_stage_display": car.get_repair_stage_display(),
                    "car_photo_file_id": car.car_photo_file_id,
                    "vin_photo_file_id": car.vin_photo_file_id,
                    "total_expenses": totals[car.id],
                    "total_expenses_by_currency": totals_by_currency[car.id],
                }
            )
        return Response(rows)


class ReportExportView(InternalAPIView):
    async def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        cars = [car async for car in Car.objects.select_related("created_by").all().order_by("status", "-created_at")]
        expenses = [
            expense
            async for expense in Expense.objects.select_related("car", "category", "employee").all().order_by("car__title", "-spent_at")
        ]

        car_ids = [car.id for car in cars]
        expense_ids = [expense.id for expense in expenses]
        totals_by_currency = defaultdict(lambda: defaultdict(lambda: Decimal("0")))
        counts = defaultdict(lambda: 0)
        for expense in expenses:
            totals_by_currency[expense.car_id][expense.currency or "BYN"] += expense.amount
            counts[expense.car_id] += 1

        car_photo_file_ids = defaultdict(list)
        car_photo_paths = defaultdict(list)
        if car_ids:
            async for photo in CarPhoto.objects.filter(car_id__in=car_ids).order_by("car_id", "created_at"):
                if photo.photo_file_id:
                    car_photo_file_ids[photo.car_id].append(photo.photo_file_id)
                if photo.image:
                    car_photo_paths[photo.car_id].append(str(photo.image))

        expense_photo_file_ids = defaultdict(list)
        expense_photo_paths = defaultdict(list)
        if expense_ids:
            async for photo in ExpensePhoto.objects.filter(expense_id__in=expense_ids).order_by("expense_id", "created_at"):
                if photo.photo_file_id:
                    expense_photo_file_ids[photo.expense_id].append(photo.photo_file_id)
                if photo.image:
                    expense_photo_paths[photo.expense_id].append(str(photo.image))

        defect_photos = []
        if car_ids:
            defect_photos = [
                photo
                async for photo in DefectPhoto.objects.select_related("car", "created_by")
                .filter(car_id__in=car_ids)
                .order_by("car__title", "created_at")
            ]

        wb = Workbook()
        ws_cars = wb.active
        ws_cars.title = "Автомобили"
        ws_expenses = wb.create_sheet("Расходы")
        ws_summary = wb.create_sheet("Итоги")
        ws_defects = wb.create_sheet("Фото дефектовки")

        def write_header(ws, headers):
            ws.append(headers)
            fill = PatternFill("solid", fgColor="D9EAF7")
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = fill

        write_header(
            ws_cars,
            [
                "ID",
                "Название",
                "Марка",
                "Модель",
                "VIN/номер",
                "Статус",
                "Этап",
                "Итого BYN",
                "Итого USD",
                "Кол-во расходов",
                "Фото авто file_id",
                "Фото авто файлы",
                "Фото VIN file_id",
                "Фото VIN файл",
                "Создан",
            ],
        )
        for car in cars:
            file_ids = car_photo_file_ids[car.id] or ([car.car_photo_file_id] if car.car_photo_file_id else [])
            paths = car_photo_paths[car.id] or ([str(car.car_photo)] if car.car_photo else [])
            ws_cars.append(
                [
                    car.id,
                    car.title,
                    car.brand,
                    car.model,
                    car.vin_or_plate or car.vin,
                    car.get_status_display(),
                    car.get_repair_stage_display(),
                    float(totals_by_currency[car.id].get("BYN", Decimal("0"))),
                    float(totals_by_currency[car.id].get("USD", Decimal("0"))),
                    counts[car.id],
                    "\n".join(file_ids),
                    "\n".join(paths),
                    car.vin_photo_file_id,
                    str(car.vin_photo) if car.vin_photo else "",
                    car.created_at.strftime("%Y-%m-%d %H:%M"),
                ]
            )

        write_header(
            ws_expenses,
            [
                "ID",
                "Автомобиль",
                "Статус авто",
                "Сумма",
                "Валюта",
                "Описание",
                "Категория",
                "Сотрудник",
                "Дата",
                "Комментарий",
                "Фото file_id",
                "Фото файлы",
            ],
        )
        for expense in expenses:
            file_ids = expense_photo_file_ids[expense.id] or ([expense.receipt_photo_file_id] if expense.receipt_photo_file_id else [])
            paths = expense_photo_paths[expense.id] or ([str(expense.receipt_photo)] if expense.receipt_photo else [])
            ws_expenses.append(
                [
                    expense.id,
                    str(expense.car),
                    expense.car.get_status_display(),
                    float(expense.amount),
                    expense.currency or "BYN",
                    expense.description,
                    expense.category.name if expense.category_id else "",
                    str(expense.employee) if expense.employee_id else "",
                    expense.spent_at.strftime("%Y-%m-%d %H:%M"),
                    expense.comment,
                    "\n".join(file_ids),
                    "\n".join(paths),
                ]
            )

        write_header(ws_summary, ["Статус", "Кол-во авто", "Кол-во расходов", "Итого BYN", "Итого USD"])
        for status_value, status_label in Car.Status.choices:
            status_cars = [car for car in cars if car.status == status_value]
            ws_summary.append(
                [
                    status_label,
                    len(status_cars),
                    sum(counts[car.id] for car in status_cars),
                    float(sum((totals_by_currency[car.id].get("BYN", Decimal("0")) for car in status_cars), Decimal("0"))),
                    float(sum((totals_by_currency[car.id].get("USD", Decimal("0")) for car in status_cars), Decimal("0"))),
                ]
            )
        ws_summary.append(
            [
                "Всего",
                len(cars),
                sum(counts.values()),
                float(sum((totals.get("BYN", Decimal("0")) for totals in totals_by_currency.values()), Decimal("0"))),
                float(sum((totals.get("USD", Decimal("0")) for totals in totals_by_currency.values()), Decimal("0"))),
            ]
        )

        write_header(ws_defects, ["ID", "Автомобиль", "Комментарий", "Добавил", "Дата", "Фото file_id", "Фото файл"])
        for photo in defect_photos:
            ws_defects.append(
                [
                    photo.id,
                    str(photo.car),
                    photo.comment,
                    str(photo.created_by) if photo.created_by_id else "",
                    photo.created_at.strftime("%Y-%m-%d %H:%M"),
                    photo.photo_file_id,
                    str(photo.image) if photo.image else "",
                ]
            )

        for ws in (ws_cars, ws_expenses, ws_summary, ws_defects):
            ws.freeze_panes = "A2"
            for column in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in column)
                ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_length + 2, 12), 45)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=auto_bot_report.xlsx"
        return response
